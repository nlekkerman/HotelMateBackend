"""
Excel Report Generator for Stock Tracker
Generates formatted Excel workbooks for stocktakes and periods
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal


def generate_stocktake_excel(stocktake):
    """
    Generate Excel workbook for a stocktake with multiple sheets.
    
    Sheets:
    1. Summary - Overview and category totals
    2. All Items - Detailed line items
    3. Category Breakdown - Items grouped by category
    4. Variance Report - Items with significant variance
    
    Args:
        stocktake: Stocktake model instance
    
    Returns:
        BytesIO: Excel file buffer
    """
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(
        start_color="4A90E2", end_color="4A90E2", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    currency_format = '€#,##0.00'
    number_format = '#,##0.00'
    
    # Borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # --- SHEET 1: SUMMARY ---
    ws_summary = wb.create_sheet("Summary")
    
    # Title
    ws_summary['A1'] = f"Stocktake Report - {stocktake.hotel.name}"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    # Period info
    row = 3
    ws_summary[f'A{row}'] = "Period:"
    ws_summary[f'B{row}'] = \
        f"{stocktake.period_start} to {stocktake.period_end}"
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Status:"
    ws_summary[f'B{row}'] = stocktake.status
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Created:"
    ws_summary[f'B{row}'] = \
        stocktake.created_at.strftime('%Y-%m-%d %H:%M')
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    if stocktake.approved_at:
        ws_summary[f'A{row}'] = "Approved:"
        ws_summary[f'B{row}'] = \
            f"{stocktake.approved_at.strftime('%Y-%m-%d %H:%M')} " \
            f"by {stocktake.approved_by}"
        ws_summary[f'A{row}'].font = bold_font
        row += 1
    
    # Summary totals
    row += 2
    lines = stocktake.lines.all()
    
    ws_summary[f'A{row}'] = "Total Items"
    ws_summary[f'B{row}'] = lines.count()
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Expected Stock Value"
    ws_summary[f'B{row}'] = \
        sum(line.expected_value for line in lines)
    ws_summary[f'B{row}'].number_format = currency_format
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Counted Stock Value"
    ws_summary[f'B{row}'] = \
        sum(line.counted_value for line in lines)
    ws_summary[f'B{row}'].number_format = currency_format
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Variance"
    ws_summary[f'B{row}'] = \
        sum(line.variance_value for line in lines)
    ws_summary[f'B{row}'].number_format = currency_format
    ws_summary[f'A{row}'].font = bold_font
    row += 2
    
    # Profitability metrics
    if stocktake.total_cogs:
        ws_summary[f'A{row}'] = "Total COGS"
        ws_summary[f'B{row}'] = float(stocktake.total_cogs)
        ws_summary[f'B{row}'].number_format = currency_format
        ws_summary[f'A{row}'].font = bold_font
        row += 1
        
        if stocktake.total_revenue:
            ws_summary[f'A{row}'] = "Total Revenue"
            ws_summary[f'B{row}'] = float(stocktake.total_revenue)
            ws_summary[f'B{row}'].number_format = currency_format
            ws_summary[f'A{row}'].font = bold_font
            row += 1
            
            ws_summary[f'A{row}'] = "Gross Profit %"
            ws_summary[f'B{row}'] = \
                f"{stocktake.gross_profit_percentage}%"
            ws_summary[f'A{row}'].font = bold_font
            row += 1
            
            ws_summary[f'A{row}'] = "Pour Cost %"
            ws_summary[f'B{row}'] = \
                f"{stocktake.pour_cost_percentage}%"
            ws_summary[f'A{row}'].font = bold_font
            row += 1
    
    # Category Totals
    row += 2
    ws_summary[f'A{row}'] = "CATEGORY BREAKDOWN"
    ws_summary[f'A{row}'].font = title_font
    row += 1
    
    # Category headers
    headers = [
        'Category', 'Opening', 'Purchases', 'Expected',
        'Counted', 'Variance', 'Variance Value'
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1
    
    category_totals = stocktake.get_category_totals()
    
    for cat_code in ['D', 'B', 'S', 'W', 'M']:
        if cat_code in category_totals:
            cat = category_totals[cat_code]
            
            ws_summary.cell(
                row=row, column=1, value=cat['category_name']
            ).border = thin_border
            ws_summary.cell(
                row=row, column=2, value=float(cat['opening_qty'])
            ).number_format = number_format
            ws_summary.cell(
                row=row, column=2
            ).border = thin_border
            ws_summary.cell(
                row=row, column=3, value=float(cat['purchases'])
            ).number_format = number_format
            ws_summary.cell(
                row=row, column=3
            ).border = thin_border
            ws_summary.cell(
                row=row, column=4, value=float(cat['expected_qty'])
            ).number_format = number_format
            ws_summary.cell(
                row=row, column=4
            ).border = thin_border
            ws_summary.cell(
                row=row, column=5, value=float(cat['counted_qty'])
            ).number_format = number_format
            ws_summary.cell(
                row=row, column=5
            ).border = thin_border
            ws_summary.cell(
                row=row, column=6, value=float(cat['variance_qty'])
            ).number_format = number_format
            ws_summary.cell(
                row=row, column=6
            ).border = thin_border
            ws_summary.cell(
                row=row, column=7, value=float(cat['variance_value'])
            ).number_format = currency_format
            ws_summary.cell(
                row=row, column=7
            ).border = thin_border
            
            row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    ws_summary.column_dimensions['F'].width = 15
    ws_summary.column_dimensions['G'].width = 18
    
    # --- SHEET 2: ALL ITEMS ---
    ws_items = wb.create_sheet("All Items")
    
    # Title
    ws_items['A1'] = "All Stock Items"
    ws_items['A1'].font = title_font
    ws_items.merge_cells('A1:L1')
    
    # Headers
    item_headers = [
        'SKU', 'Name', 'Category', 'Opening Qty', 'Purchases',
        'Expected Qty', 'Counted Full', 'Counted Partial',
        'Counted Qty', 'Variance Qty', 'Expected Value',
        'Counted Value', 'Variance Value'
    ]
    
    row = 3
    for col_idx, header in enumerate(item_headers, start=1):
        cell = ws_items.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    row += 1
    
    # Data rows
    for line in lines:
        ws_items.cell(
            row=row, column=1, value=line.item.sku
        ).border = thin_border
        ws_items.cell(
            row=row, column=2, value=line.item.name
        ).border = thin_border
        ws_items.cell(
            row=row, column=3, value=line.item.category.code
        ).border = thin_border
        ws_items.cell(
            row=row, column=4, value=float(line.opening_qty)
        ).number_format = number_format
        ws_items.cell(
            row=row, column=4
        ).border = thin_border
        ws_items.cell(
            row=row, column=5, value=float(line.purchases)
        ).number_format = number_format
        ws_items.cell(
            row=row, column=5
        ).border = thin_border
        ws_items.cell(
            row=row, column=6, value=float(line.expected_qty)
        ).number_format = number_format
        ws_items.cell(
            row=row, column=6
        ).border = thin_border
        ws_items.cell(
            row=row, column=7, value=float(line.counted_full_units)
        ).number_format = number_format
        ws_items.cell(
            row=row, column=7
        ).border = thin_border
        ws_items.cell(
            row=row, column=8, value=float(line.counted_partial_units)
        ).number_format = number_format
        ws_items.cell(
            row=row, column=8
        ).border = thin_border
        ws_items.cell(
            row=row, column=9, value=float(line.counted_qty)
        ).number_format = number_format
        ws_items.cell(
            row=row, column=9
        ).border = thin_border
        ws_items.cell(
            row=row, column=10, value=float(line.variance_qty)
        ).number_format = number_format
        ws_items.cell(
            row=row, column=10
        ).border = thin_border
        ws_items.cell(
            row=row, column=11, value=float(line.expected_value)
        ).number_format = currency_format
        ws_items.cell(
            row=row, column=11
        ).border = thin_border
        ws_items.cell(
            row=row, column=12, value=float(line.counted_value)
        ).number_format = currency_format
        ws_items.cell(
            row=row, column=12
        ).border = thin_border
        ws_items.cell(
            row=row, column=13, value=float(line.variance_value)
        ).number_format = currency_format
        ws_items.cell(
            row=row, column=13
        ).border = thin_border
        
        row += 1
    
    # Adjust column widths
    for col_idx in range(1, 14):
        ws_items.column_dimensions[
            get_column_letter(col_idx)
        ].width = 15
    ws_items.column_dimensions['B'].width = 30  # Name column
    
    # --- SHEET 3: VARIANCE REPORT ---
    ws_variance = wb.create_sheet("Variance Report")
    
    # Title
    ws_variance['A1'] = "Variance Report (Items with Variance)"
    ws_variance['A1'].font = title_font
    ws_variance.merge_cells('A1:H1')
    
    # Headers
    variance_headers = [
        'SKU', 'Name', 'Category', 'Expected Qty',
        'Counted Qty', 'Variance Qty', 'Variance Value', '% Variance'
    ]
    
    row = 3
    for col_idx, header in enumerate(variance_headers, start=1):
        cell = ws_variance.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1
    
    # Only include items with variance
    variance_lines = [
        line for line in lines if abs(line.variance_qty) > 0.01
    ]
    
    for line in variance_lines:
        variance_pct = 0
        if line.expected_qty > 0:
            variance_pct = (
                line.variance_qty / line.expected_qty * 100
            )
        
        ws_variance.cell(
            row=row, column=1, value=line.item.sku
        ).border = thin_border
        ws_variance.cell(
            row=row, column=2, value=line.item.name
        ).border = thin_border
        ws_variance.cell(
            row=row, column=3, value=line.item.category.code
        ).border = thin_border
        ws_variance.cell(
            row=row, column=4, value=float(line.expected_qty)
        ).number_format = number_format
        ws_variance.cell(
            row=row, column=4
        ).border = thin_border
        ws_variance.cell(
            row=row, column=5, value=float(line.counted_qty)
        ).number_format = number_format
        ws_variance.cell(
            row=row, column=5
        ).border = thin_border
        ws_variance.cell(
            row=row, column=6, value=float(line.variance_qty)
        ).number_format = number_format
        ws_variance.cell(
            row=row, column=6
        ).border = thin_border
        ws_variance.cell(
            row=row, column=7, value=float(line.variance_value)
        ).number_format = currency_format
        ws_variance.cell(
            row=row, column=7
        ).border = thin_border
        ws_variance.cell(
            row=row, column=8, value=float(variance_pct)
        ).number_format = '0.00"%"'
        ws_variance.cell(
            row=row, column=8
        ).border = thin_border
        
        row += 1
    
    # Adjust column widths
    ws_variance.column_dimensions['A'].width = 15
    ws_variance.column_dimensions['B'].width = 30
    ws_variance.column_dimensions['C'].width = 12
    ws_variance.column_dimensions['D'].width = 15
    ws_variance.column_dimensions['E'].width = 15
    ws_variance.column_dimensions['F'].width = 15
    ws_variance.column_dimensions['G'].width = 18
    ws_variance.column_dimensions['H'].width = 15
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_period_excel(period, include_cocktails=True):
    """
    Generate Excel workbook for a period with multiple sheets.
    
    Sheets:
    1. Summary - Period overview
    2. Stock Snapshots - All items with closing stock
    3. Category Analysis - Breakdown by category
    4. Sales Analysis - Sales data if available
    
    Args:
        period: StockPeriod model instance
        include_cocktails: Include cocktail sales data
    
    Returns:
        BytesIO: Excel file buffer
    """
    wb = Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(
        start_color="4A90E2", end_color="4A90E2", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    currency_format = '€#,##0.00'
    number_format = '#,##0.00'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get snapshots
    from stock_tracker.models import StockSnapshot
    snapshots = StockSnapshot.objects.filter(
        period=period
    ).select_related('item', 'item__category')
    
    # --- SHEET 1: SUMMARY ---
    ws_summary = wb.create_sheet("Summary")
    
    # Title
    ws_summary['A1'] = f"Period Report - {period.hotel.name}"
    ws_summary['A1'].font = title_font
    ws_summary.merge_cells('A1:D1')
    
    # Period info
    row = 3
    ws_summary[f'A{row}'] = "Period:"
    ws_summary[f'B{row}'] = period.period_name
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Dates:"
    ws_summary[f'B{row}'] = \
        f"{period.start_date} to {period.end_date}"
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Type:"
    ws_summary[f'B{row}'] = period.get_period_type_display()
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Status:"
    ws_summary[f'B{row}'] = "Closed" if period.is_closed else "Open"
    ws_summary[f'A{row}'].font = bold_font
    row += 2
    
    # Summary totals
    total_value = sum(s.closing_stock_value for s in snapshots)
    
    ws_summary[f'A{row}'] = "Total Items"
    ws_summary[f'B{row}'] = snapshots.count()
    ws_summary[f'A{row}'].font = bold_font
    row += 1
    
    ws_summary[f'A{row}'] = "Closing Stock Value"
    ws_summary[f'B{row}'] = float(total_value)
    ws_summary[f'B{row}'].number_format = currency_format
    ws_summary[f'A{row}'].font = bold_font
    row += 2
    
    # Cocktail data
    if include_cocktails:
        ws_summary[f'A{row}'] = "Cocktail Sales Revenue"
        ws_summary[f'B{row}'] = float(period.cocktail_revenue)
        ws_summary[f'B{row}'].number_format = currency_format
        ws_summary[f'A{row}'].font = bold_font
        row += 1
        
        ws_summary[f'A{row}'] = "Cocktails Made"
        ws_summary[f'B{row}'] = period.cocktail_quantity
        ws_summary[f'A{row}'].font = bold_font
        row += 2
    
    # Category breakdown
    ws_summary[f'A{row}'] = "CATEGORY BREAKDOWN"
    ws_summary[f'A{row}'].font = title_font
    row += 1
    
    cat_headers = ['Category', 'Items', 'Value', '% of Total']
    for col_idx, header in enumerate(cat_headers, start=1):
        cell = ws_summary.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1
    
    # Calculate category totals
    category_data = {}
    for snapshot in snapshots:
        cat_code = snapshot.item.category.code
        if cat_code not in category_data:
            category_data[cat_code] = {
                'name': snapshot.item.category.name,
                'value': Decimal('0.00'),
                'items': 0
            }
        category_data[cat_code]['value'] += \
            snapshot.closing_stock_value
        category_data[cat_code]['items'] += 1
    
    for cat_code in ['D', 'B', 'S', 'W', 'M']:
        if cat_code in category_data:
            cat = category_data[cat_code]
            percentage = (
                cat['value'] / total_value * 100
            ) if total_value > 0 else 0
            
            ws_summary.cell(
                row=row, column=1, value=cat['name']
            ).border = thin_border
            ws_summary.cell(
                row=row, column=2, value=cat['items']
            ).border = thin_border
            ws_summary.cell(
                row=row, column=3, value=float(cat['value'])
            ).number_format = currency_format
            ws_summary.cell(
                row=row, column=3
            ).border = thin_border
            ws_summary.cell(
                row=row, column=4, value=float(percentage)
            ).number_format = '0.0"%"'
            ws_summary.cell(
                row=row, column=4
            ).border = thin_border
            
            row += 1
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 20
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 15
    
    # --- SHEET 2: STOCK SNAPSHOTS ---
    ws_snapshots = wb.create_sheet("Stock Snapshots")
    
    # Title
    ws_snapshots['A1'] = "Stock Snapshots"
    ws_snapshots['A1'].font = title_font
    ws_snapshots.merge_cells('A1:H1')
    
    # Headers
    snapshot_headers = [
        'SKU', 'Name', 'Category', 'Size',
        'Full Units', 'Partial Units',
        'Total Servings', 'Value'
    ]
    
    row = 3
    for col_idx, header in enumerate(snapshot_headers, start=1):
        cell = ws_snapshots.cell(
            row=row, column=col_idx, value=header
        )
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1
    
    # Data rows
    for snapshot in snapshots:
        ws_snapshots.cell(
            row=row, column=1, value=snapshot.item.sku
        ).border = thin_border
        ws_snapshots.cell(
            row=row, column=2, value=snapshot.item.name
        ).border = thin_border
        ws_snapshots.cell(
            row=row, column=3, value=snapshot.item.category.code
        ).border = thin_border
        ws_snapshots.cell(
            row=row, column=4, value=snapshot.item.size
        ).border = thin_border
        ws_snapshots.cell(
            row=row, column=5,
            value=float(snapshot.closing_full_units)
        ).number_format = number_format
        ws_snapshots.cell(
            row=row, column=5
        ).border = thin_border
        ws_snapshots.cell(
            row=row, column=6,
            value=float(snapshot.closing_partial_units)
        ).number_format = number_format
        ws_snapshots.cell(
            row=row, column=6
        ).border = thin_border
        ws_snapshots.cell(
            row=row, column=7,
            value=float(snapshot.total_servings)
        ).number_format = number_format
        ws_snapshots.cell(
            row=row, column=7
        ).border = thin_border
        ws_snapshots.cell(
            row=row, column=8,
            value=float(snapshot.closing_stock_value)
        ).number_format = currency_format
        ws_snapshots.cell(
            row=row, column=8
        ).border = thin_border
        
        row += 1
    
    # Adjust column widths
    ws_snapshots.column_dimensions['A'].width = 15
    ws_snapshots.column_dimensions['B'].width = 35
    ws_snapshots.column_dimensions['C'].width = 12
    ws_snapshots.column_dimensions['D'].width = 15
    ws_snapshots.column_dimensions['E'].width = 12
    ws_snapshots.column_dimensions['F'].width = 15
    ws_snapshots.column_dimensions['G'].width = 15
    ws_snapshots.column_dimensions['H'].width = 18
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
